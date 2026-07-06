"""Edge-path tests: ingest warnings (duplicate/orphan rows), validation error
branches (no units, out-of-range GPS, missing area/operator), parse_area
rejects, and the no-farms guard in run(). Complements test_eudr_agent.py,
which covers the primary branches via the synthetic S13 fixture.
"""
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from eudr_agent import (  # noqa: E402
    Farm, Unit, decimal_places, ingest, parse_area, run, summarize, validate,
)


def _sheet(wb, prefix):
    for name in wb.sheetnames:
        if name.strip().startswith(prefix):
            return wb[name]
    raise AssertionError(f"fixture is missing sheet {prefix}")


# ---------------------------------------------------------------- parse_area

@pytest.mark.parametrize("raw", ["-1", "-0.5", None, "abc", ""])
def test_parse_area_rejects_negative_and_garbage(raw):
    assert parse_area(raw) is None


def test_parse_area_accepts_zero_and_comma_decimal():
    assert parse_area("0") == pytest.approx(0.0)
    assert parse_area("1,6") == pytest.approx(1.6)


# ---------------------------------------------------------------- ingest edges

def test_ingest_duplicate_farm_id_keeps_first_and_warns(s13_file):
    wb = openpyxl.load_workbook(s13_file)
    farm_sheet = _sheet(wb, "1.")
    # Same id as the first data row, different village: must NOT overwrite.
    farm_sheet.append(["TF - AA-1", "nd", "ELSEWHERE", "Sud-Kivu", "Butamu", "9.9"])
    wb.save(s13_file)

    farms = {f.farm_id: f for f in ingest(s13_file)}
    assert len(farms) == 6  # duplicate did not add a farm
    f1 = farms["TF-AA-1"]
    assert f1.village == "Kabamba"  # first occurrence kept
    assert any("Duplicate farm_id" in w for w in f1.warnings)


def test_ingest_orphan_unit_rows_skipped_with_warning(s13_file, capsys):
    wb = openpyxl.load_workbook(s13_file)
    unit_sheet = _sheet(wb, "3.")
    unit_sheet.append(["TF - ZZ-99", "TF - ZZ-99", "0.5", "-2.1998402", "28.8658507"])
    wb.save(s13_file)

    farms = ingest(s13_file)
    assert len(farms) == 6
    assert sum(len(f.units) for f in farms) == 7  # orphan unit not attached
    assert "1 unit rows reference unknown farm ids" in capsys.readouterr().err


# ---------------------------------------------------------------- validate edges

def _minimal_farm(**overrides) -> Farm:
    farm = Farm(
        farm_id="T-1",
        area_ha=1.0,
        operator_name="Test Person",
        inspector="INSPECTOR",
        inspection_year="2025",
        units=[Unit(unit_id="T-1", area_ha=1.0,
                    lat_raw="-2.1998402", lon_raw="28.8658507")],
    )
    for key, value in overrides.items():
        setattr(farm, key, value)
    return farm


def test_validate_farm_without_units_is_gps_error():
    farm = _minimal_farm(units=[])
    validate([farm])
    assert any(e.startswith("GPS: no farm units") for e in farm.errors)
    assert not farm.gps_ready


@pytest.mark.parametrize("lat,lon", [
    ("95.1234567", "28.8658507"),    # latitude beyond +/-90
    ("-2.1998402", "-181.1234567"),  # longitude beyond +/-180
])
def test_validate_out_of_range_coordinates(lat, lon):
    farm = _minimal_farm(units=[Unit("T-1", 1.0, lat, lon)])
    validate([farm])
    assert any("coordinates out of range" in e for e in farm.errors)
    assert not farm.gps_ready


def test_validate_missing_area_and_operator():
    farm = _minimal_farm(area_ha=None, operator_name="")
    validate([farm])
    assert "DATA: total farm area missing" in farm.errors
    assert "DATA: operator name missing" in farm.errors
    # GPS itself is fine — DATA errors must not disturb gps_ready.
    assert farm.gps_ready
    assert not farm.compliant


# ------------------------------------------------------------ decimal_places

@pytest.mark.parametrize("raw,expected", [
    ("+1.234567", 6),    # explicit plus sign
    (".1234567", 7),     # leading-dot decimal
    ("-2.1998402", 7),
    ("28", 0),
    (-1.5, 1),           # float fallback via repr() — documented lower bound
])
def test_decimal_places_tolerates_sign_and_bare_dot_forms(raw, expected):
    assert decimal_places(raw) == expected


# ------------------------------------------------------------ score flooring

def test_score_floors_so_only_full_compliance_reads_100(tmp_path):
    """9996/10000 compliant must NOT round up to 100 — the public badge's
    'Ready' state keys off scorePercent === 100."""
    source = tmp_path / "registry.xlsx"
    source.write_bytes(b"fake-registry-bytes")
    bad = _minimal_farm(farm_id="T-BAD")
    bad.errors.append("DATA: operator name missing")
    farms = [_minimal_farm(farm_id=f"T-{i}") for i in range(1999)] + [bad]
    score = summarize(farms, source, "Test", "Café")["eudrCompliance"]["scorePercent"]
    assert score == 99.9  # floor(99.95) at 1 decimal, never 100.0

    all_good = [_minimal_farm(farm_id=f"T-{i}") for i in range(2000)]
    assert summarize(all_good, source, "Test", "Café")["eudrCompliance"]["scorePercent"] == 100.0


# ------------------------------------------------------------ ragged rows

def test_ingest_survives_ragged_rows(s13_file):
    """read_only sheets can yield short tuples; a row holding only a farm id
    must ingest (with missing-data errors later), not crash the run."""
    wb = openpyxl.load_workbook(s13_file)
    _sheet(wb, "1.").append(["TF - RG-1"])
    _sheet(wb, "2.").append(["TF - RG-1"])
    _sheet(wb, "3.").append(["TF - RG-1"])
    wb.save(s13_file)

    farms = {f.farm_id: f for f in ingest(s13_file)}
    assert "TF-RG-1" in farms
    ragged = farms["TF-RG-1"]
    assert ragged.area_ha is None and ragged.operator_name == ""
    validate([ragged])
    assert not ragged.compliant  # flagged, not crashed


# ---------------------------------------------------------------- run guard

def test_run_raises_on_empty_registry(tmp_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for prefix in ("1. vide", "2. vide", "3. vide"):
        sheet = wb.create_sheet(prefix)
        sheet.append(["TITRE"])
        sheet.append(["entêtes"])
    path = tmp_path / "empty_s13.xlsx"
    wb.save(path)
    with pytest.raises(ValueError, match="No farms parsed"):
        run(path, tmp_path / "out", "Test", "Café")
