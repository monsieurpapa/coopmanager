"""Synthetic RA S13 fixture — mirrors the real template's structure with
invented farms covering every validation branch. The real registry (farmer PII)
is gitignored and must NEVER be used as test data.
"""
import pytest
import openpyxl

SHEET_FARM = "1. l'exploitation agricole"
SHEET_CROP = "2. Produit agricole certifié"
SHEET_UNIT = "3. Unité agricole"


def _pad(row, width):
    return row + [""] * (width - len(row))


@pytest.fixture
def s13_file(tmp_path):
    """Six farms exercising: clean pass, short precision, missing coords,
    >4 ha polygon flag, duplicate coordinates, multi-unit farm, and a
    missing-inspection error. Farm ids intentionally use the messy
    'XX - YY-n' spacing of the real template.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    farm = wb.create_sheet(SHEET_FARM)
    farm.append(_pad(["EXPLOITATION AGRICOLE"], 27))
    farm.append(_pad([f"{i}. col" for i in range(1, 28)], 27))
    #        id            natl  village    district    region   area  type  units crops first     last      tel  nid  genre    yob    (owner 16-20)      workers x2   inspector      yr    mo   day
    rows = [
        ["TF - AA-1", "nd", "Kabamba", "Sud-Kivu", "Butamu", "0.4", "P", "1", "1", "Alice",  "Test",  "243", "0", "Femme", "1970", "", "", "", "", "", "1", "0", "INSPECTOR ONE", "2025", "12", "9"],
        ["TF - AA-2", "nd", "Kabamba", "Sud-Kivu", "Butamu", "0.7", "P", "1", "1", "Bob",    "Test",  "243", "0", "Autre", "1980", "", "", "", "", "", "1", "0", "INSPECTOR ONE", "2025", "12", "9"],
        ["TF - AA-3", "nd", "Mabingu", "Sud-Kivu", "Butamu", "1.1", "P", "1", "1", "Carla",  "Test",  "243", "0", "Femme", "1975", "", "", "", "", "", "1", "0", "INSPECTOR ONE", "2025", "12", "9"],
        ["TF - AA-4", "nd", "Mabingu", "Sud-Kivu", "Butamu", "5.5", "P", "1", "1", "Denis",  "Test",  "243", "0", "Autre", "1965", "", "", "", "", "", "1", "0", "INSPECTOR TWO", "2025", "12", "9"],
        ["TF - AA-5", "nd", "Kasheke", "Sud-Kivu", "Butamu", "0.9", "P", "1", "1", "Esther", "Test",  "243", "0", "Femme", "1990", "", "", "", "", "", "1", "0", "INSPECTOR TWO", "2025", "12", "9"],
        ["TF - AA-6", "nd", "Kasheke", "Sud-Kivu", "Butamu", "1.6", "P", "2", "1", "Felix",  "Test",  "243", "0", "Autre", "1985", "", "", "", "", "", "1", "0", "", "", "", ""],
    ]
    for r in rows:
        farm.append(_pad(r, 27))

    crop = wb.create_sheet(SHEET_CROP)
    crop.append(_pad(["CULTURE CERTIFIÉE"], 11))
    crop.append(_pad([f"{i}, col" for i in range(1, 12)], 11))
    for fid, area, harvest in [
        ("TF - AA-1", "0.4", "228"), ("TF - AA-2", "0.7", "399"),
        ("TF - AA-3", "1.1", "500"), ("TF - AA-4", "5.5", "2500"),
        ("TF - AA-5", "0.9", "410"), ("TF - AA-6", "1.6", "700"),
    ]:
        crop.append(_pad([fid, "Café", "Arabica", area, harvest, harvest, harvest], 11))

    unit = wb.create_sheet(SHEET_UNIT)
    unit.append(_pad(["INFORMATIONS UNITÉS"], 9))
    unit.append(_pad([f"{i}. col" for i in range(1, 10)], 9))
    units = [
        # AA-1: clean — 7 decimals both axes
        ["TF - AA-1", "TF - AA-1", "0.4", "-2.1998402", "28.8658507"],
        # AA-2: precision below the 6-decimal floor (4 decimals)
        ["TF - AA-2", "TF - AA-2", "0.7", "-2.1999", "28.8656"],
        # AA-3: missing longitude
        ["TF - AA-3", "TF - AA-3", "1.1", "-2.1998028", None],
        # AA-4: clean GPS but farm >4 ha (polygon warning, not GPS error)
        ["TF - AA-4", "TF - AA-4", "5.5", "-2.1970469", "28.8633455"],
        # AA-5: duplicates AA-1's coordinates exactly
        ["TF - AA-5", "TF - AA-5", "0.9", "-2.1998402", "28.8658507"],
        # AA-6: two units, both clean (multi-unit farm; missing inspection in sheet 1)
        ["TF - AA-6", "TF - AA-6-a", "0.8", "-2.1963638", "28.8607419"],
        ["TF - AA-6", "TF - AA-6-b", "0.8", "-2.1956983", "28.8611682"],
    ]
    for u in units:
        unit.append(_pad(u, 9))

    path = tmp_path / "synthetic_s13.xlsx"
    wb.save(path)
    return path
