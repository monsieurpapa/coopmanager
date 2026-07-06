#!/usr/bin/env python3
"""EUDR geolocation-readiness agent for RA Annex S13 cooperative registries.

Reads a Rainforest Alliance S13 Group Member Registry (.xlsx, French template),
validates plot geolocation against EUDR requirements, and emits:

  eudr_compliance.json   summary for the CongoFarmers /cooperatives doc
                         (paste into the admin EUDR box; includes provenance)
  validation_report.json per-farm ERROR/WARNING detail — ADMIN EYES ONLY
                         (contains farm ids + coordinates)
  buyer_document.geojson data-minimized FeatureCollection for the buyer's
                         due-diligence filing (plot geolocation, farm_id,
                         area, product — NO farmer names/gender/age)
  buyer_summary.pdf      one-page human-readable summary, no PII

EUDR notes encoded here:
  - Art. 2(28): geolocation of plots <4 ha may be a single point with at
    least SIX decimal digits of latitude/longitude. Precision is checked on
    the raw cell text (the registry stores coordinates as strings); floats
    fall back to repr(), the shortest round-tripping representation.
  - Plots >4 ha require polygon boundaries. The S13 template carries no
    polygons, so any farm over 4 ha sets oversizedFarmsMissingPolygon=true
    and is listed in the validation report for follow-up.
  - The output attests GEOLOCATION READINESS ONLY — the S13 registry holds
    no forest-cover data, so no deforestation-free claim is made anywhere.

Usage:
  python eudr_agent.py --input path/to/GMR.xlsx --output reports/ \
      [--coop-name "Maendeleo"] [--product "Coffee (Arabica)"]
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

SCRIPT_VERSION = "1.0.0"
MIN_DECIMALS = 6          # EUDR Art. 2(28) floor for point geolocation
POLYGON_THRESHOLD_HA = 4.0

# Sheet name prefixes in the RA S13 French template. Matched by prefix so a
# renamed accent or trailing text does not break ingestion.
SHEET_FARM = "1."
SHEET_CROP = "2."
SHEET_UNIT = "3."
HEADER_ROWS = 2  # both a title row and a column-header row precede data


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------

def normalize_id(raw) -> str:
    """Standardize farm ids: 'KB - AM-1 ' -> 'KB-AM-1' (join key + display)."""
    if raw is None:
        return ""
    return re.sub(r"\s*-\s*", "-", str(raw).strip())


def decimal_places(raw) -> int:
    """Count decimal digits as entered.

    Strings are authoritative (the S13 stores coordinates as text). Floats
    use repr(), the shortest string that round-trips the double — a faithful
    lower bound on the digits originally entered.
    """
    if raw is None:
        return 0
    text = str(raw).strip().replace(",", ".") if isinstance(raw, str) else repr(raw)
    match = re.match(r"^[+-]?\d*\.(\d+)$", text)  # tolerate '+1.5' and '.5' forms
    return len(match.group(1)) if match else 0


def parse_coord(raw) -> float | None:
    if raw is None:
        return None
    try:
        return float(str(raw).strip().replace(",", "."))
    except ValueError:
        return None


def parse_area(raw) -> float | None:
    try:
        value = float(str(raw).strip().replace(",", "."))
        return value if value >= 0 else None
    except (TypeError, ValueError):
        return None


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

@dataclass
class Unit:
    unit_id: str
    area_ha: float | None
    lat_raw: object
    lon_raw: object

    @property
    def lat(self) -> float | None:
        return parse_coord(self.lat_raw)

    @property
    def lon(self) -> float | None:
        return parse_coord(self.lon_raw)


@dataclass
class Farm:
    farm_id: str
    village: str = ""
    district: str = ""
    area_ha: float | None = None
    operator_name: str = ""
    gender: str = ""
    inspector: str = ""
    inspection_year: object = None
    product: str = ""
    variety: str = ""
    harvest_kg: float | None = None
    units: list[Unit] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def gps_ready(self) -> bool:
        """Every unit has an in-range point with >= MIN_DECIMALS precision."""
        if not self.units:
            return False
        return all(not e.startswith("GPS") for e in self.errors)

    @property
    def compliant(self) -> bool:
        return not self.errors


# ---------------------------------------------------------------------------
# Ingestion
# ---------------------------------------------------------------------------

def _cell(row: tuple, index: int):
    """Ragged-row-safe indexing: read_only worksheets can yield short tuples
    when the sheet's dimension record is off, and row[n] would raise."""
    return row[index] if len(row) > index else None


def _find_sheet(workbook, prefix: str):
    for name in workbook.sheetnames:
        if name.strip().startswith(prefix):
            return workbook[name]
    raise ValueError(
        f"Sheet starting with '{prefix}' not found. "
        f"Sheets present: {workbook.sheetnames}. Is this an RA S13 file?"
    )


def ingest(path: Path) -> list[Farm]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        farm_sheet = _find_sheet(workbook, SHEET_FARM)
        crop_sheet = _find_sheet(workbook, SHEET_CROP)
        unit_sheet = _find_sheet(workbook, SHEET_UNIT)

        farms: dict[str, Farm] = {}
        for row in farm_sheet.iter_rows(min_row=HEADER_ROWS + 1, values_only=True):
            farm_id = normalize_id(_cell(row, 0))
            if not farm_id:
                continue  # trailing empty rows
            if farm_id in farms:
                farms[farm_id].warnings.append("Duplicate farm_id in sheet 1 — first occurrence kept")
                continue
            first = str(_cell(row, 9)).strip() if _cell(row, 9) else ""
            last = str(_cell(row, 10)).strip() if _cell(row, 10) else ""
            farms[farm_id] = Farm(
                farm_id=farm_id,
                village=str(_cell(row, 2)).strip() if _cell(row, 2) else "",
                district=str(_cell(row, 3)).strip() if _cell(row, 3) else "",
                area_ha=parse_area(_cell(row, 5)),
                operator_name=f"{first} {last}".strip(),
                gender=str(_cell(row, 13)).strip() if _cell(row, 13) else "",
                inspector=str(_cell(row, 22)).strip() if _cell(row, 22) else "",
                inspection_year=_cell(row, 23),
            )

        for row in crop_sheet.iter_rows(min_row=HEADER_ROWS + 1, values_only=True):
            farm_id = normalize_id(_cell(row, 0))
            farm = farms.get(farm_id)
            if not farm:
                continue
            farm.product = str(_cell(row, 1)).strip() if _cell(row, 1) else ""
            farm.variety = str(_cell(row, 2)).strip() if _cell(row, 2) else ""
            farm.harvest_kg = parse_area(_cell(row, 4))

        orphan_units = 0
        for row in unit_sheet.iter_rows(min_row=HEADER_ROWS + 1, values_only=True):
            farm_id = normalize_id(_cell(row, 0))
            if not farm_id:
                continue
            farm = farms.get(farm_id)
            if not farm:
                orphan_units += 1
                continue
            farm.units.append(Unit(
                unit_id=normalize_id(_cell(row, 1)) or farm_id,
                area_ha=parse_area(_cell(row, 2)),
                lat_raw=_cell(row, 3),
                lon_raw=_cell(row, 4),
            ))
        if orphan_units:
            print(f"WARNING: {orphan_units} unit rows reference unknown farm ids (skipped)",
                  file=sys.stderr)

        return list(farms.values())
    finally:
        workbook.close()


# ---------------------------------------------------------------------------
# Validation & scoring
# ---------------------------------------------------------------------------

def validate(farms: list[Farm]) -> None:
    seen_coords: dict[tuple[float, float], str] = {}
    for farm in farms:
        if not farm.units:
            farm.errors.append("GPS: no farm units / coordinates recorded")
        for unit in farm.units:
            lat, lon = unit.lat, unit.lon
            if lat is None or lon is None:
                farm.errors.append(f"GPS: unit {unit.unit_id} missing or unparseable coordinates")
                continue
            if not (-90 <= lat <= 90) or not (-180 <= lon <= 180):
                farm.errors.append(f"GPS: unit {unit.unit_id} coordinates out of range ({lat}, {lon})")
                continue
            lat_dp = decimal_places(unit.lat_raw)
            lon_dp = decimal_places(unit.lon_raw)
            if lat_dp < MIN_DECIMALS or lon_dp < MIN_DECIMALS:
                farm.errors.append(
                    f"GPS: unit {unit.unit_id} precision below EUDR Art. 2(28) "
                    f"floor of {MIN_DECIMALS} decimals (lat {lat_dp}dp, lon {lon_dp}dp)"
                )
            key = (round(lat, 7), round(lon, 7))
            if key in seen_coords and seen_coords[key] != farm.farm_id:
                farm.warnings.append(
                    f"GPS: unit {unit.unit_id} duplicates coordinates of farm "
                    f"{seen_coords[key]} — verify data entry"
                )
            else:
                seen_coords[key] = farm.farm_id

        if farm.area_ha is None:
            farm.errors.append("DATA: total farm area missing")
        if not farm.operator_name:
            farm.errors.append("DATA: operator name missing")
        if not farm.inspector or not farm.inspection_year:
            farm.errors.append("DATA: internal inspection record incomplete")
        if farm.area_ha is not None and farm.area_ha > POLYGON_THRESHOLD_HA:
            farm.warnings.append(
                f"POLYGON: farm is {farm.area_ha} ha (> {POLYGON_THRESHOLD_HA} ha) — "
                "EUDR requires polygon boundaries, which the S13 template cannot carry"
            )


def summarize(farms: list[Farm], source: Path, coop_name: str, product: str) -> dict:
    total = len(farms)
    gps_ready = sum(1 for f in farms if f.gps_ready)
    compliant = sum(1 for f in farms if f.compliant)
    oversized = [f for f in farms
                 if f.area_ha is not None and f.area_ha > POLYGON_THRESHOLD_HA]
    digest = hashlib.sha256(source.read_bytes()).hexdigest()
    return {
        "eudrCompliance": {
            # floor, not round: 9996/10000 compliant must NOT display as 100
            # — the badge's "Ready" state keys off scorePercent === 100, so
            # 100.0 may only appear when every farm is compliant.
            "scorePercent": math.floor(1000.0 * compliant / total) / 10.0 if total else 0.0,
            "totalFarms": total,
            "farmsWithGps": gps_ready,
            "oversizedFarmsMissingPolygon": bool(oversized),
            "computedAt": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
            "sourceFileHash": digest,
            "sourceFileName": source.name,
            "scriptVersion": SCRIPT_VERSION,
        },
        "context": {
            "cooperative": coop_name,
            "product": product,
            "claimScope": "geolocation readiness only — no deforestation-free assessment",
            "minDecimalsRequired": MIN_DECIMALS,
            "farmsOver4haCount": len(oversized),
            "totalAreaHa": round(sum(f.area_ha for f in farms if f.area_ha is not None), 2),
        },
    }


# ---------------------------------------------------------------------------
# Outputs
# ---------------------------------------------------------------------------

def emit_validation_report(farms: list[Farm], out_dir: Path) -> Path:
    report = {
        "generatedAt": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "note": "ADMIN EYES ONLY — contains farm ids and coordinate diagnostics. Do not send to buyers.",
        "farms": [
            {
                "farm_id": f.farm_id,
                "village": f.village,
                "status": "OK" if f.compliant and not f.warnings
                          else ("ERROR" if f.errors else "WARNING"),
                "errors": f.errors,
                "warnings": f.warnings,
            }
            for f in farms if f.errors or f.warnings
        ],
    }
    path = out_dir / "validation_report.json"
    path.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    return path


def emit_buyer_geojson(farms: list[Farm], out_dir: Path, product: str) -> Path:
    """Data-minimized plot geolocation for the buyer's due-diligence filing.

    Properties carry ONLY what a due-diligence statement needs: plot id, area,
    product. Farmer names, gender, and ages deliberately never enter this file.
    Coordinates are emitted as [lon, lat] with 6-decimal rounding, matching the
    EU Information System (TRACES) GeoJSON conventions for point geolocation.
    """
    features = []
    for farm in farms:
        for unit in farm.units:
            if unit.lat is None or unit.lon is None:
                continue
            features.append({
                "type": "Feature",
                "geometry": {
                    "type": "Point",
                    "coordinates": [round(unit.lon, 6), round(unit.lat, 6)],
                },
                "properties": {
                    "ProducerName": farm.farm_id,   # id, not a person's name
                    "ProductionPlace": unit.unit_id,
                    "Area": unit.area_ha,
                    "product": product,
                },
            })
    doc = {"type": "FeatureCollection", "features": features}
    path = out_dir / "buyer_document.geojson"
    path.write_text(json.dumps(doc, indent=2), encoding="utf-8")
    return path


def emit_buyer_pdf(summary: dict, out_dir: Path) -> Path:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    c = summary["eudrCompliance"]
    ctx = summary["context"]
    lines = [
        ("EUDR Geolocation-Readiness Summary", 16, True),
        (f"Cooperative: {ctx['cooperative']}", 12, False),
        (f"Product: {ctx['product']}", 12, False),
        ("", 10, False),
        (f"Farms in registry: {c['totalFarms']}    "
         f"Total area: {ctx['totalAreaHa']} ha", 11, False),
        (f"Farms geolocation-ready (point, >= {ctx['minDecimalsRequired']} decimals): "
         f"{c['farmsWithGps']}", 11, False),
        (f"Overall readiness score: {c['scorePercent']}%", 13, True),
        (f"Farms over 4 ha needing polygon boundaries: {ctx['farmsOver4haCount']}", 11, False),
        ("", 10, False),
        ("Scope: this document attests geolocation readiness only.", 10, False),
        ("It makes no deforestation-free assessment (EUDR Art. 3 due", 10, False),
        ("diligence remains the operator's obligation).", 10, False),
        ("", 10, False),
        (f"Computed: {c['computedAt']}   Script v{c['scriptVersion']}", 9, False),
        (f"Source registry SHA-256: {c['sourceFileHash'][:32]}...", 9, False),
        ("Plot geolocations: see accompanying buyer_document.geojson", 9, False),
    ]
    fig = plt.figure(figsize=(8.27, 11.69))  # A4 portrait
    y = 0.93
    for text, size, bold in lines:
        fig.text(0.08, y, text, fontsize=size,
                 fontweight="bold" if bold else "normal", family="sans-serif")
        y -= 0.045 if size >= 13 else 0.035
    path = out_dir / "buyer_summary.pdf"
    fig.savefig(path, format="pdf")
    plt.close(fig)
    return path


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def run(input_path: Path, out_dir: Path, coop_name: str, product: str) -> dict:
    out_dir.mkdir(parents=True, exist_ok=True)
    farms = ingest(input_path)
    if not farms:
        raise ValueError("No farms parsed — check the input file structure.")
    validate(farms)
    summary = summarize(farms, input_path, coop_name, product)

    compliance_path = out_dir / "eudr_compliance.json"
    compliance_path.write_text(json.dumps(summary, indent=2, ensure_ascii=False),
                               encoding="utf-8")
    emit_validation_report(farms, out_dir)
    emit_buyer_geojson(farms, out_dir, product)
    emit_buyer_pdf(summary, out_dir)

    c = summary["eudrCompliance"]
    print(f"Parsed {c['totalFarms']} farms | GPS-ready {c['farmsWithGps']} | "
          f"score {c['scorePercent']}% | >4ha missing polygon: "
          f"{c['oversizedFarmsMissingPolygon']} "
          f"({summary['context']['farmsOver4haCount']} farms)")
    print(f"Outputs written to {out_dir}/")
    print("Paste the 'eudrCompliance' object from eudr_compliance.json into the "
          "admin EUDR box in the app.")
    return summary


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--input", required=True, type=Path,
                        help="RA S13 registry .xlsx")
    parser.add_argument("--output", default=Path("reports"), type=Path,
                        help="output directory (default: reports/)")
    parser.add_argument("--coop-name", default="Maendeleo")
    parser.add_argument("--product", default="Coffee (Arabica)")
    args = parser.parse_args()
    run(args.input, args.output, args.coop_name, args.product)


if __name__ == "__main__":
    main()
